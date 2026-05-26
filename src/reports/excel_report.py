"""엑셀 분석 보고서 생성

사용법:
    python -m src.reports.excel_report --region 11680 --months 12 --output report.xlsx
"""
from __future__ import annotations
import argparse
import json
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config.settings import REPORT_DIR
from src.database.repository import fetch_trades_df, fetch_rents_df
from src.analysis.price_trend import monthly_summary, apt_summary, yoy_change
from src.analysis.gap_analysis import gap_table
from src.analysis.yield_calc import rental_yield
from src.utils.logger import get_logger

log = get_logger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)


def _write_df(ws, df: pd.DataFrame, start_row: int = 1):
    if df.empty:
        ws.cell(row=start_row, column=1, value="(데이터 없음)")
        return start_row + 2
    for c_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=str(col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for r_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for c_idx, val in enumerate(row, 1):
            if pd.isna(val):
                v = ""
            elif isinstance(val, (pd.Timestamp,)):
                v = val.strftime("%Y-%m-%d")
            else:
                v = val
            ws.cell(row=r_idx, column=c_idx, value=v)
    for c_idx, col in enumerate(df.columns, 1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()[:200]]
        )
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 40)
    return start_row + len(df) + 2


def _add_line_chart(ws, df: pd.DataFrame, x_col: str, y_cols: list[str],
                    title: str, anchor: str, start_row: int):
    if df.empty:
        return
    cols = list(df.columns)
    nrows = len(df)
    chart = LineChart()
    chart.title = title
    chart.height = 9
    chart.width = 18
    for y in y_cols:
        if y not in cols:
            continue
        col_idx = cols.index(y) + 1
        ref = Reference(ws, min_col=col_idx, max_col=col_idx,
                        min_row=start_row, max_row=start_row + nrows)
        chart.add_data(ref, titles_from_data=True)
    x_idx = cols.index(x_col) + 1
    cats = Reference(ws, min_col=x_idx, max_col=x_idx,
                     min_row=start_row + 1, max_row=start_row + nrows)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def build_report(region_code: str, months: int, output: Path):
    today = date.today()
    yr, mo = today.year, today.month - months
    while mo <= 0:
        mo += 12
        yr -= 1
    date_from = date(yr, mo, 1)

    log.info("기간: %s ~ %s, 지역: %s", date_from, today, region_code)

    df_t = fetch_trades_df(region_code=region_code, date_from=date_from)
    df_r = fetch_rents_df(region_code=region_code, date_from=date_from)

    log.info("trades=%d, rents=%d", len(df_t), len(df_r))

    monthly = yoy_change(monthly_summary(df_t))
    apts = apt_summary(df_t, top=50)
    gap = gap_table(df_t, df_r, area_tol=5.0, months=min(months, 6))
    yld = rental_yield(df_t, df_r, area_tol=5.0, months=months)

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "요약"
    ws_sum["A1"] = "부동산 분석 보고서"
    ws_sum["A1"].font = TITLE_FONT
    ws_sum["A2"] = f"지역코드: {region_code}"
    ws_sum["A3"] = f"기간: {date_from} ~ {today}"
    ws_sum["A4"] = f"매매 건수: {len(df_t):,} / 전월세 건수: {len(df_r):,}"

    if not monthly.empty:
        ws_sum["A6"] = "월별 추이"
        ws_sum["A6"].font = Font(bold=True)
        _write_df(ws_sum, monthly, start_row=7)
        _add_line_chart(ws_sum, monthly, "ym",
                        ["avg_price", "median_price"],
                        "월별 평균/중위 매매가 (만원)", "H7", start_row=7)
        _add_line_chart(ws_sum, monthly, "ym",
                        ["avg_ppp"],
                        "평당가 추이 (만원/평)", "H30", start_row=7)

    ws_apt = wb.create_sheet("단지별")
    _write_df(ws_apt, apts)

    ws_gap = wb.create_sheet("매매-전세 갭")
    _write_df(ws_gap, gap)

    ws_yld = wb.create_sheet("임대수익률")
    _write_df(ws_yld, yld)

    ws_raw_t = wb.create_sheet("매매원본")
    _write_df(ws_raw_t, df_t.head(5000))

    ws_raw_r = wb.create_sheet("전월세원본")
    _write_df(ws_raw_r, df_r.head(5000))

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    log.info("보고서 저장: %s", output)
    return output


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--region", required=True, help="법정동 시군구 5자리 (예: 11680)")
    ap.add_argument("--months", type=int, default=12)
    ap.add_argument("--output", default=None)
    args = ap.parse_args()

    out = Path(args.output) if args.output else REPORT_DIR / f"report_{args.region}_{date.today():%Y%m%d}.xlsx"
    build_report(args.region, args.months, out)
    print(f"OK: {out}")


if __name__ == "__main__":
    main()
