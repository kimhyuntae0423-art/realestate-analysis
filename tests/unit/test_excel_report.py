"""src/reports/excel_report.py — 엑셀 보고서 생성 검증."""
from __future__ import annotations

from datetime import date

import pandas as pd
from openpyxl import Workbook, load_workbook

from src.reports.excel_report import build_report, _write_df
from src.database.repository import upsert_trades, upsert_rents


def test_write_df_empty_dataframe_writes_placeholder():
    wb = Workbook()
    ws = wb.active
    next_row = _write_df(ws, pd.DataFrame(), start_row=1)
    assert ws.cell(row=1, column=1).value == "(데이터 없음)"
    assert next_row == 3


def test_write_df_writes_header_and_rows():
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame({"a": [1, 2], "b": ["x", "y"]})
    next_row = _write_df(ws, df, start_row=1)
    assert ws.cell(row=1, column=1).value == "a"
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=3, column=2).value == "y"
    assert next_row == 1 + len(df) + 2


def _trade_row(d, region="11680", apt="A", amount=100000):
    return {"region_code": region, "deal_date": d, "deal_year": d.year,
            "deal_month": d.month, "deal_day": d.day, "apt_name": apt,
            "area_m2": 84.9, "deal_amount": amount, "price_per_pyeong": 6000}


def _rent_row(d, region="11680", apt="A", deposit=70000):
    return {"region_code": region, "deal_date": d, "deal_year": d.year,
            "deal_month": d.month, "deal_day": d.day, "apt_name": apt,
            "area_m2": 84.9, "deposit": deposit, "monthly_rent": 0}


def test_build_report_creates_workbook_with_expected_sheets(tmp_path):
    upsert_trades([_trade_row(date(2025, 6, 1)), _trade_row(date(2025, 7, 1), amount=110000)])
    upsert_rents([_rent_row(date(2025, 6, 15))])

    out_path = tmp_path / "report.xlsx"
    result = build_report("11680", months=12, output=out_path)

    assert result == out_path
    assert out_path.exists()
    wb = load_workbook(out_path)
    assert wb.sheetnames == ["요약", "단지별", "매매-전세 갭", "임대수익률", "매매원본", "전월세원본"]
    assert wb["요약"]["A2"].value == "지역코드: 11680"


def test_build_report_handles_no_data(tmp_path):
    out_path = tmp_path / "empty_report.xlsx"
    result = build_report("99999", months=12, output=out_path)
    assert result.exists()
    wb = load_workbook(out_path)
    assert wb["단지별"]["A1"].value == "(데이터 없음)"
